"""
Build / refresh /data/cyfoods-ledger.xlsx from the canonical purchase data
in this script. Single source of truth for the historical purchase ledger
that mirrors what's seeded into the Postgres DB.

To add new purchases later: extend PURCHASES below and re-run:
    python3 data/build_ledger.py
"""

from __future__ import annotations

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

KEG_SIZE_L = 25

# (date, supplier, kegs, price/keg, logistics/keg, note)
PURCHASES = [
    ("2026-01-19", "Tomike from Ondo", 5,  60_000, 0,     "Source quoted total only (300k); price/keg derived."),
    ("2026-02-10", "Tomike from Ondo", 7,  45_000, 0,     "Stated total 315k confirmed (315/7 = 45k/keg)."),
    ("2026-02-11", "Tomike from Ondo", 10, 48_000, 0,     ""),
    ("2026-02-16", "Tomike from Ondo", 20, 41_500, 4_000, "Logistics 4k/keg from later message — only 20-keg load on record."),
    ("2026-03-17", "Daniela PNC Tropical Foods",          5,  50_500, 0,     ""),
    ("2026-03-19", "Daniela PNC Tropical Foods",          10, 41_500, 6_000, "Logistics 6k/keg — assumed applies to most recent 10-keg load (two such loads exist)."),
]

OUT = Path(__file__).resolve().parent / "cyfoods-ledger.xlsx"

# ---------- styles ----------
HEADER_FILL = PatternFill("solid", fgColor="1F4F3F")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="Calibri", size=16, bold=True, color="1F4F3F")
NOTE_FONT   = Font(name="Calibri", size=9, italic=True, color="6B6B6B")
TOTAL_FILL  = PatternFill("solid", fgColor="EAF2EE")
TOTAL_FONT  = Font(name="Calibri", size=11, bold=True)

THIN = Side(border_style="thin", color="C7CFCB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CURRENCY_FMT = '"₦"#,##0'
NUMBER_FMT   = '#,##0'
DECIMAL_FMT  = '#,##0.00'


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER


def style_body(cell, fmt=None):
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BORDER
    if fmt:
        cell.number_format = fmt


def autosize(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------- Sheet 1: Purchases ----------

def build_purchases(ws):
    headers = [
        "Date",
        "Supplier",
        "Kegs",
        "Keg size (L)",
        "Price per keg (₦)",
        "Logistics per keg (₦)",
        "Total logistics (₦)",
        "Total cost (₦)",
        "Cost per litre (₦)",
        "Status",
        "Notes / assumptions",
    ]

    ws.cell(row=1, column=1, value="CYFoods — Purchase Ledger").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=2, column=1, value="Source: Mary's WhatsApp records · Mirrored in /prisma/seed.ts").font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16

    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=col, value=h)
        style_header(c)
    ws.row_dimensions[4].height = 32

    total_kegs = 0
    total_logistics = 0
    total_cost = 0
    total_litres = 0

    start = 5
    for i, (date, supplier, kegs, ppk, lpk, note) in enumerate(PURCHASES):
        row = start + i
        keg_logistics = kegs * lpk
        cost = kegs * ppk + keg_logistics
        litres = kegs * KEG_SIZE_L
        cpl = cost / litres if litres else 0

        total_kegs += kegs
        total_logistics += keg_logistics
        total_cost += cost
        total_litres += litres

        cells = [
            (1, date,      None),
            (2, supplier,  None),
            (3, kegs,      NUMBER_FMT),
            (4, KEG_SIZE_L, NUMBER_FMT),
            (5, ppk,       CURRENCY_FMT),
            (6, lpk,       CURRENCY_FMT),
            (7, keg_logistics, CURRENCY_FMT),
            (8, cost,      CURRENCY_FMT),
            (9, round(cpl, 2), DECIMAL_FMT),
            (10, "ACCEPTED", None),
            (11, note,     None),
        ]
        for col, value, fmt in cells:
            c = ws.cell(row=row, column=col, value=value)
            style_body(c, fmt)

    # Totals row
    trow = start + len(PURCHASES)
    label = ws.cell(row=trow, column=1, value="TOTAL")
    label.font = TOTAL_FONT
    label.fill = TOTAL_FILL
    label.border = BORDER

    for col in range(2, len(headers) + 1):
        c = ws.cell(row=trow, column=col)
        c.fill = TOTAL_FILL
        c.font = TOTAL_FONT
        c.border = BORDER

    ws.cell(row=trow, column=3, value=total_kegs).number_format = NUMBER_FMT
    ws.cell(row=trow, column=7, value=total_logistics).number_format = CURRENCY_FMT
    ws.cell(row=trow, column=8, value=total_cost).number_format = CURRENCY_FMT
    avg_cpl = total_cost / total_litres if total_litres else 0
    ws.cell(row=trow, column=9, value=round(avg_cpl, 2)).number_format = DECIMAL_FMT

    autosize(ws, [13, 22, 8, 12, 18, 20, 18, 18, 16, 12, 50])
    ws.freeze_panes = "A5"


# ---------- Sheet 2: Summary ----------

def build_summary(ws):
    rows = [
        ("CYFoods — Stock & Purchase Summary", None, "title"),
        ("As mirrored in production database", None, "note"),
        ("", None, None),
        ("Metric", "Value", "header"),
    ]

    total_kegs = sum(p[2] for p in PURCHASES)
    total_litres = total_kegs * KEG_SIZE_L
    total_cost = sum(p[2] * p[3] + p[2] * p[4] for p in PURCHASES)
    total_logistics = sum(p[2] * p[4] for p in PURCHASES)
    avg_cpl = total_cost / total_litres if total_litres else 0
    avg_kpk = sum(p[3] for p in PURCHASES) / len(PURCHASES)

    metrics = [
        ("Total purchase events", len(PURCHASES), NUMBER_FMT),
        ("Total kegs purchased",  total_kegs,     NUMBER_FMT),
        ("Total litres",          total_litres,   NUMBER_FMT),
        ("Total cost",            total_cost,     CURRENCY_FMT),
        ("Total logistics paid",  total_logistics, CURRENCY_FMT),
        ("Average cost per litre", round(avg_cpl, 2), DECIMAL_FMT),
        ("Average price per keg (unweighted)", round(avg_kpk, 0), CURRENCY_FMT),
        ("", "", None),
        ("Kegs currently in stock (no sales yet)", total_kegs, NUMBER_FMT),
        ("Litres currently in stock", total_litres, NUMBER_FMT),
        ("Stock value (cost basis)", total_cost, CURRENCY_FMT),
    ]

    # Title
    t = ws.cell(row=1, column=1, value=rows[0][0])
    t.font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    n = ws.cell(row=2, column=1, value=rows[1][0])
    n.font = NOTE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)

    # Header
    h1 = ws.cell(row=4, column=1, value="Metric")
    h2 = ws.cell(row=4, column=2, value="Value")
    style_header(h1)
    style_header(h2)
    ws.row_dimensions[4].height = 28

    # Body
    for i, (label, value, fmt) in enumerate(metrics, start=5):
        c1 = ws.cell(row=i, column=1, value=label)
        c2 = ws.cell(row=i, column=2, value=value)
        style_body(c1)
        style_body(c2, fmt)

    autosize(ws, [42, 22])

    # Suppliers breakdown
    sup_start = 5 + len(metrics) + 2
    ws.cell(row=sup_start, column=1, value="By supplier").font = TOTAL_FONT
    ws.cell(row=sup_start + 1, column=1, value="Supplier").font = HEADER_FONT
    ws.cell(row=sup_start + 1, column=1).fill = HEADER_FILL
    ws.cell(row=sup_start + 1, column=2, value="Kegs").font = HEADER_FONT
    ws.cell(row=sup_start + 1, column=2).fill = HEADER_FILL
    ws.cell(row=sup_start + 1, column=1).border = BORDER
    ws.cell(row=sup_start + 1, column=2).border = BORDER

    by_supplier: dict[str, dict[str, float]] = {}
    for date, supplier, kegs, ppk, lpk, note in PURCHASES:
        d = by_supplier.setdefault(supplier, {"kegs": 0, "cost": 0})
        d["kegs"] += kegs
        d["cost"] += kegs * ppk + kegs * lpk

    for i, (sup, d) in enumerate(sorted(by_supplier.items()), start=sup_start + 2):
        c1 = ws.cell(row=i, column=1, value=sup)
        c2 = ws.cell(row=i, column=2, value=int(d["kegs"]))
        style_body(c1)
        style_body(c2, NUMBER_FMT)


# ---------- Sheet 3: Open issues / questions for Mary ----------

def build_open_questions(ws):
    ws.cell(row=1, column=1, value="Open questions / data hygiene").font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    rows = [
        ("Q1", "Logistics: '20 rubbers @ 4k each' was applied to the 20-keg Feb 16 load (Tomike). '10 rubbers @ 6k each' was applied to Mar 19 (Daniela PNC). Confirm — could it have applied to Feb 11 instead?"),
        ("Q2", "Were any of these kegs sold, packed, or repackaged before today? (Currently we assume all 57 kegs are still on hand.)"),
        ("Q3", "Any other purchases between Jan 19 and Mar 19, or after Mar 19, that aren't on this list?"),
    ]

    h1 = ws.cell(row=3, column=1, value="#")
    h2 = ws.cell(row=3, column=2, value="Question")
    style_header(h1)
    style_header(h2)
    ws.row_dimensions[3].height = 28

    for i, (n, q) in enumerate(rows, start=4):
        c1 = ws.cell(row=i, column=1, value=n)
        c2 = ws.cell(row=i, column=2, value=q)
        style_body(c1)
        style_body(c2)
        ws.row_dimensions[i].height = 38

    autosize(ws, [6, 110])


def main():
    wb = Workbook()
    purchases_ws = wb.active
    purchases_ws.title = "Purchases"
    build_purchases(purchases_ws)

    summary_ws = wb.create_sheet("Summary")
    build_summary(summary_ws)

    questions_ws = wb.create_sheet("Open questions")
    build_open_questions(questions_ws)

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
